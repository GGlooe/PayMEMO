from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, Form, UploadFile, File, Depends, HTTPException, Query, Header
from fastapi.responses import RedirectResponse, JSONResponse, Response, HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from app.auth import (
    check_auth, verify_credentials, create_session, get_csrf_token,
    require_csrf, delete_session, get_session, change_password, USERNAME
)
from app import payments, storage, notifications
import threading
import time
import csv
import io
from datetime import date


@asynccontextmanager
async def lifespan(app: FastAPI):
    t = threading.Thread(target=scheduler_loop, daemon=True)
    t.start()
    yield

app = FastAPI(lifespan=lifespan)

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
templates.env.globals["money_text"] = payments.money_text


def render(request: Request, name: str, **context):
    context.setdefault("user", None)
    context.setdefault("csrf_token", get_csrf_token(request))
    return templates.TemplateResponse(request=request, name=name, context=context)


# ---------- Фоновый поток (только desktop-уведомления Windows) ----------
def scheduler_loop():
    while True:
        notifications.send_desktop_notifications()
        time.sleep(60)


# ---------- Service Worker + иконка ----------
@app.get("/sw.js")
async def serve_sw():
    return FileResponse("static/sw.js")


@app.get("/static/icon.png")
async def icon():
    return Response(
        content=b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\xdac\xf8\x00\x00\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82',
        media_type="image/png"
    )


# ---------- Подготовка данных ----------
def get_status(payment):
    if payments.is_paid(payment):
        return "оплачен", "status-paid"
    if payment.get("paid_amount", 0) > 0:
        return "частично оплачен", "status-partial"
    try:
        due = payments.parse_date(payment["pay_date"])
        today = date.today()
        if due < today:
            return "просрочен", "status-unpaid"
        elif due == today:
            return "сегодня", "status-soon"
    except:
        pass
    return "ожидает", "status-unpaid"


def prepare_payment(p):
    p = dict(p)
    p["status_text"], p["status_class"] = get_status(p)
    p["formatted_date"] = payments.format_date(p.get("pay_date", ""))
    p["formatted_amount"] = payments.money_text(p.get("amount", 0), p.get("currency", "RUB"))
    p.setdefault("comment", p.get("description", ""))
    return p


# ---------- Главная ----------
@app.get("/")
async def root(request: Request, user: str = Depends(check_auth)):
    all_p = payments.load_payments()
    active = [prepare_payment(p) for p in all_p if not payments.is_paid(p) and not p.get("archived", False)]
    archived = payments.archived_payments()
    settings = storage.load_settings()
    monthly_report = payments.build_monthly_report()

    return render(
        request, "index.html",
        user=user,
        payments=active,
        archive=archived,
        settings=settings,
        monthly_report=monthly_report,
    )


# ---------- Авторизация ----------
@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = None):
    return render(request, "login.html", error=error)


@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    if verify_credentials(username, password):
        session_id, _ = create_session()
        response = RedirectResponse(url="/", status_code=302)
        response.set_cookie(key="session_id", value=session_id, httponly=True, samesite="lax")
        return response
    return render(request, "login.html", error="Неверный логин или пароль")


@app.post("/logout")
async def logout(request: Request, csrf: bool = Depends(require_csrf)):
    session = get_session(request)
    if session:
        delete_session(session.get("id"))
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("session_id")
    return response


@app.get("/change-password", response_class=HTMLResponse)
async def change_password_page(request: Request, user: str = Depends(check_auth)):
    return render(request, "change_password.html", user=user)


@app.post("/change-password")
async def change_password_post(request: Request, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf),
                               current_password: str = Form(...), new_password: str = Form(...)):
    if not verify_credentials(USERNAME, current_password):
        return render(request, "change_password.html", user=user, error="Текущий пароль неверен")
    change_password(new_password)
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("session_id")
    return response


# ---------- Платежи ----------
@app.get("/payments", response_class=HTMLResponse)
async def payments_page(request: Request, filter: str = Query("all"), user: str = Depends(check_auth)):
    all_p = payments.load_payments()
    if filter == "soon":
        filtered = [p for p in all_p if payments.is_soon(p) and not payments.is_paid(p)]
    elif filter == "urgent":
        filtered = [p for p in all_p if payments.is_overdue(p) and not payments.is_paid(p)]
    elif filter == "paid":
        filtered = [p for p in all_p if payments.is_paid(p)]
    else:
        filtered = [p for p in all_p if not payments.is_paid(p) and not p.get("archived", False)]

    urgent_count = len([p for p in all_p if payments.is_overdue(p) and not payments.is_paid(p)])
    soon_count = len([p for p in all_p if payments.is_soon(p) and not payments.is_paid(p)])
    paid_count = len([p for p in all_p if payments.is_paid(p)])

    prepared = [prepare_payment(p) for p in filtered]

    return render(
        request, "payments.html",
        user=user,
        payments=prepared,
        urgent_count=urgent_count,
        soon_count=soon_count,
        paid_count=paid_count,
    )


# ---------- Добавление ----------
@app.get("/add", response_class=HTMLResponse)
async def add_form(request: Request, user: str = Depends(check_auth)):
    return render(request, "add.html", user=user)


@app.post("/add")
async def add_payment(request: Request, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf),
                      bank_name: str = Form(""), amount: float = Form(...),
                      paid_amount: float = Form(0), currency: str = Form("RUB"),
                      pay_date: str = Form(...), comment: str = Form("")):
    payments.add_payment(name=bank_name, pay_date=pay_date, amount=amount,
                         paid_amount=paid_amount, currency=currency, description=comment)
    storage.add_history_entry("add", bank_name, f"Добавлен платёж на сумму {amount} {currency}")
    return RedirectResponse(url="/payments", status_code=302)


# ---------- Редактирование ----------
@app.get("/edit/{payment_id}", response_class=HTMLResponse)
async def edit_form(request: Request, payment_id: int, user: str = Depends(check_auth)):
    p = payments.find_payment(payment_id)
    if not p:
        raise HTTPException(status_code=404, detail="Платёж не найден")
    return render(request, "edit.html", user=user, payment=prepare_payment(p))


@app.post("/edit/{payment_id}")
async def edit_payment(request: Request, payment_id: int, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf),
                       bank_name: str = Form(""), amount: float = Form(...),
                       paid_amount: float = Form(0), currency: str = Form("RUB"),
                       pay_date: str = Form(...), comment: str = Form("")):
    payments.update_payment(payment_id, name=bank_name, pay_date=pay_date, amount=amount,
                            paid_amount=paid_amount, currency=currency, description=comment)
    storage.add_history_entry("edit", bank_name, "Изменён платёж")
    return RedirectResponse(url="/payments", status_code=302)


# ---------- Оплата / частичная / удаление ----------
@app.post("/pay/{payment_id}")
async def mark_paid(payment_id: int, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf)):
    p = payments.find_payment(payment_id)
    if p:
        payments.mark_paid(payment_id)
        storage.add_history_entry("pay", p.get("bank_name", ""), "Платёж полностью оплачен")
    return RedirectResponse(url="/payments", status_code=302)


@app.post("/partial/{payment_id}")
async def mark_partial(payment_id: int, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf),
                       paid_amount: float = Form(...)):
    p = payments.find_payment(payment_id)
    if p:
        payments.mark_partial(payment_id, paid_amount)
        storage.add_history_entry("partial", p.get("bank_name", ""), f"Частичная оплата {paid_amount}")
    return RedirectResponse(url="/payments", status_code=302)


@app.post("/delete/{payment_id}")
async def delete_payment_post(payment_id: int, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf)):
    p = payments.find_payment(payment_id)
    if p:
        name = p.get("bank_name", "")
        payments.delete_payment(payment_id)
        storage.add_history_entry("delete", name, "Платёж удалён")
    return RedirectResponse(url="/payments", status_code=302)


# ---------- Архив ----------
@app.get("/archive", response_class=HTMLResponse)
async def archive_page(request: Request, user: str = Depends(check_auth)):
    archived = [prepare_payment(p) for p in payments.archived_payments()]
    return render(request, "archive.html", user=user, payments=archived)


# ---------- История ----------
@app.get("/history", response_class=HTMLResponse)
async def history_page(request: Request, user: str = Depends(check_auth)):
    hist = storage.load_history()
    return render(request, "history.html", user=user, history=hist)


# ---------- Отчёты ----------
@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, user: str = Depends(check_auth)):
    report_data = payments.build_monthly_report()
    settings = storage.load_settings()
    return render(request, "reports.html", user=user, report=report_data, settings=settings)


# ---------- Настройки ----------
@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, user: str = Depends(check_auth)):
    settings = storage.load_settings()
    return render(request, "settings.html", user=user, settings=settings)


@app.post("/settings")
async def settings_update(request: Request, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf)):
    form = await request.form()
    new_settings = {
        "remind_days_before": int(form.get("remind_days_before", 3)),
        "remind_time_1": form.get("remind_time_1", "09:00"),
        "remind_time_2": form.get("remind_time_2", "18:00"),
        "notify_due_day": "notify_due_day" in form,
        "notify_month_end": "notify_month_end" in form,
        "browser_notifications": "browser_notifications" in form,
        "check_interval_minutes": int(form.get("check_interval_minutes", 30)),
        "notifications_enabled": "notifications_enabled" in form,
    }
    storage.save_settings(new_settings)
    return RedirectResponse(url="/settings", status_code=302)


# ---------- Импорт/Экспорт ----------
@app.get("/import", response_class=HTMLResponse)
async def import_page(request: Request, user: str = Depends(check_auth)):
    return render(request, "import.html", user=user)


@app.post("/import")
async def import_payments(request: Request, user: str = Depends(check_auth), csrf: bool = Depends(require_csrf),
                          file: UploadFile = File(...)):
    if file:
        try:
            payments.import_from_file(file.file, file.filename)
            storage.add_history_entry("import", "", f"Импортированы платежи из {file.filename}")
        except RuntimeError as e:
            return render(request, "import.html", user=user, error=str(e))
    return RedirectResponse(url="/payments", status_code=302)


@app.get("/export/payments")
async def export_payments(user: str = Depends(check_auth)):
    xlsx_data = payments.export_payments_xlsx()
    if xlsx_data:
        return Response(content=xlsx_data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=payments.xlsx"})
    csv_data = payments.export_payments_csv()
    return Response(content=csv_data, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=payments.csv"})


@app.get("/export/archive")
async def export_archive(user: str = Depends(check_auth)):
    xlsx_data = payments.export_archive_xlsx()
    if xlsx_data:
        return Response(content=xlsx_data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=archive.xlsx"})
    csv_data = payments.export_archive_csv()
    return Response(content=csv_data, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=archive.csv"})


@app.get("/history/export")
async def export_history(user: str = Depends(check_auth)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        headers = ["Время", "Действие", "Платёж", "Детали"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        hist = storage.load_history()
        for row in hist:
            ws.append([
                row.get("ts", ""), row.get("action", ""), row.get("bank_name", ""), row.get("details", "")
            ])
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=history.xlsx"})
    except ImportError:
        hist = storage.load_history()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Время", "Действие", "Платёж", "Детали"])
        for row in hist:
            writer.writerow([row.get("ts", ""), row.get("action", ""), row.get("bank_name", ""), row.get("details", "")])
        return Response(content=output.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": "attachment; filename=history.csv"})


@app.get("/export/reports")
async def export_reports(user: str = Depends(check_auth)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        headers = ["Месяц", "Начислено", "Оплачено", "Всего", "Оплачено", "Частично", "Не оплачено"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        report_data = payments.build_monthly_report()
        for row in report_data:
            ws.append([
                row["month"], row["total_amount"], row["total_paid"], row["count"],
                row["paid_count"], row["partial_count"], row["unpaid_count"]
            ])
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=reports.xlsx"})
    except ImportError:
        return JSONResponse({"error": "Установите openpyxl: pip install openpyxl"}, status_code=500)


# ---------- Уведомления ----------
@app.post("/notifications/test")
async def test_notification(user: str = Depends(check_auth), csrf: bool = Depends(require_csrf)):
    result = notifications.get_pending_notifications(test=True)
    return JSONResponse(result)


@app.get("/api/notifications/poll")
async def poll_notifications(user: str = Depends(check_auth)):
    result = notifications.get_pending_notifications()
    return JSONResponse(result)