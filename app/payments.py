from pathlib import Path
import json
import csv
import io
from datetime import datetime, date, timedelta

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
PAYMENTS_FILE = DATA_DIR / "payments.json"

CURRENCY_SYMBOLS = {"RUB": "₽", "USD": "$", "EUR": "€"}
MONTHS_RU = {1: "янв", 2: "фев", 3: "мар", 4: "апр", 5: "мая", 6: "июн",
             7: "июл", 8: "авг", 9: "сен", 10: "окт", 11: "ноя", 12: "дек"}


def load_json(path, default):
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return default
    return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_payments():
    data = load_json(PAYMENTS_FILE, [])
    # Всегда сортируем по дате платежа (ближайшие сверху)
    data.sort(key=lambda p: p.get("pay_date", ""))
    return data


def save_payments(data):
    save_json(PAYMENTS_FILE, data)


def parse_date(date_str):
    return datetime.strptime(date_str, "%Y-%m-%d").date()


def next_payment_id():
    all_p = load_payments()
    return max([p["id"] for p in all_p], default=0) + 1


def find_payment(payment_id: int):
    for p in load_payments():
        if p["id"] == payment_id:
            return p
    return None


def is_paid(p):
    return p.get("paid_amount", 0) >= p.get("amount", 0)


def is_overdue(p, today=None):
    if is_paid(p) or p.get("archived", False):
        return False
    try:
        d = parse_date(p["pay_date"])
        today = today or date.today()
        return d < today
    except:
        return False


def is_soon(p, days_before=None):
    if days_before is None:
        from app import storage
        days_before = int(storage.load_settings().get("remind_days_before", 3))
    if is_paid(p) or p.get("archived", False):
        return False
    try:
        d = parse_date(p["pay_date"])
        today = date.today()
        delta = (d - today).days
        return 0 <= delta <= int(days_before)
    except:
        return False


def archived_payments():
    return [p for p in load_payments() if p.get("archived", False) or is_paid(p)]


def active_payments():
    return [p for p in load_payments() if not is_paid(p) and not p.get("archived", False)]


def add_payment(name, pay_date, amount, paid_amount=0, currency="RUB", description="", category="", recurring=""):
    all_p = load_payments()
    new_id = next_payment_id()
    new_p = {
        "id": new_id,
        "name": name,
        "bank_name": name,
        "pay_date": pay_date,
        "amount": float(amount),
        "paid_amount": float(paid_amount),
        "currency": currency,
        "description": description,
        "comment": description,
        "category": category,
        "recurring": recurring,
        "archived": False
    }
    all_p.append(new_p)
    save_payments(all_p)
    return new_id


def update_payment(payment_id, name=None, pay_date=None, amount=None,
                   paid_amount=None, currency=None, description=None,
                   category=None, recurring=None, paid=False, archived=False):
    all_p = load_payments()
    for p in all_p:
        if p["id"] == payment_id:
            if name is not None:
                p["name"] = name
                p["bank_name"] = name
            if pay_date is not None:
                p["pay_date"] = pay_date
            if amount is not None:
                p["amount"] = float(amount)
            if paid_amount is not None:
                p["paid_amount"] = float(paid_amount)
            if currency is not None:
                p["currency"] = currency
            if description is not None:
                p["description"] = description
                p["comment"] = description
            if category is not None:
                p["category"] = category
            if recurring is not None:
                p["recurring"] = recurring
            if paid:
                p["paid_amount"] = p["amount"]
            p["archived"] = archived
            break
    save_payments(all_p)


def delete_payment(payment_id):
    all_p = load_payments()
    all_p = [p for p in all_p if p["id"] != payment_id]
    save_payments(all_p)


def mark_paid(payment_id):
    all_p = load_payments()
    for p in all_p:
        if p["id"] == payment_id:
            p["paid_amount"] = p["amount"]
            break
    save_payments(all_p)


def mark_partial(payment_id, amount):
    all_p = load_payments()
    for p in all_p:
        if p["id"] == payment_id:
            p["paid_amount"] = min(float(amount), p["amount"])
            break
    save_payments(all_p)


def _import_rows(rows):
    all_p = load_payments()
    for row in rows:
        try:
            name = row.get("name") or row.get("bank_name") or row.get("Банк") or "Без названия"
            pay_date = row.get("pay_date") or row.get("Дата") or date.today().isoformat()
            amount = float(row.get("amount") or row.get("Сумма") or 0)
            currency = row.get("currency") or row.get("Валюта") or "RUB"
            comment = row.get("comment") or row.get("description") or row.get("Комментарий") or ""
            new_id = max([p["id"] for p in all_p], default=0) + 1
            all_p.append({
                "id": new_id,
                "name": name,
                "bank_name": name,
                "pay_date": pay_date,
                "amount": amount,
                "paid_amount": 0,
                "currency": currency,
                "description": comment,
                "comment": comment,
                "category": "",
                "recurring": "",
                "archived": False
            })
        except Exception:
            continue
    save_payments(all_p)


def import_from_file(file_obj, filename: str):
    if filename.lower().endswith(('.xlsx', '.xls')):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_obj)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for h, v in zip(headers, row):
                    if h:
                        row_dict[h] = v
                rows.append(row_dict)
            _import_rows(rows)
        except ImportError:
            raise RuntimeError("Для импорта Excel установите: pip install openpyxl")
    else:
        content = file_obj.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
        _import_rows(list(reader))


def export_payments_csv():
    all_p = load_payments()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Название", "Дата", "Сумма", "Оплачено", "Валюта", "Категория", "Комментарий", "Архив"])
    for p in all_p:
        writer.writerow([p["id"], p.get("name",""), p.get("pay_date",""), p.get("amount",0),
                         p.get("paid_amount",0), p.get("currency",""), p.get("category",""),
                         p.get("description",""), "Да" if p.get("archived", False) else "Нет"])
    return output.getvalue()


def export_archive_csv():
    archived = archived_payments()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Название", "Дата", "Сумма", "Оплачено", "Валюта", "Категория", "Комментарий"])
    for p in archived:
        writer.writerow([p["id"], p.get("name",""), p.get("pay_date",""), p.get("amount",0),
                         p.get("paid_amount",0), p.get("currency",""), p.get("category",""),
                         p.get("description","")])
    return output.getvalue()


def _build_xlsx(rows, headers):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return None
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_payments_xlsx():
    all_p = load_payments()
    headers = ["ID", "Название", "Дата", "Сумма", "Оплачено", "Валюта", "Категория", "Комментарий", "Архив"]
    rows = []
    for p in all_p:
        rows.append([
            p["id"], p.get("name",""), p.get("pay_date",""), p.get("amount",0),
            p.get("paid_amount",0), p.get("currency",""), p.get("category",""),
            p.get("description",""), "Да" if p.get("archived", False) else "Нет"
        ])
    return _build_xlsx(rows, headers)


def export_archive_xlsx():
    archived = archived_payments()
    headers = ["ID", "Название", "Дата", "Сумма", "Оплачено", "Валюта", "Категория", "Комментарий"]
    rows = []
    for p in archived:
        rows.append([
            p["id"], p.get("name",""), p.get("pay_date",""), p.get("amount",0),
            p.get("paid_amount",0), p.get("currency",""), p.get("category",""),
            p.get("description","")
        ])
    return _build_xlsx(rows, headers)


def build_monthly_report():
    report = {}
    for p in load_payments():
        try:
            dt = datetime.strptime(p["pay_date"], "%Y-%m-%d")
        except:
            continue
        key = f"{dt.year}-{dt.month:02d}"
        if key not in report:
            report[key] = {
                "month": key,
                "total_amount": 0.0,
                "total_paid": 0.0,
                "count": 0,
                "paid_count": 0,
                "partial_count": 0,
                "unpaid_count": 0,
            }
        amount = float(p.get("amount", 0))
        paid = float(p.get("paid_amount", 0))
        report[key]["total_amount"] += amount
        report[key]["total_paid"] += paid
        report[key]["count"] += 1
        if paid <= 0:
            report[key]["unpaid_count"] += 1
        elif paid < amount:
            report[key]["partial_count"] += 1
        else:
            report[key]["paid_count"] += 1
    return [report[k] for k in sorted(report.keys())]


def format_date(date_str: str) -> str:
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return f"{dt.day} {MONTHS_RU[dt.month]} {dt.year}"
    except:
        return date_str


def money_text(value: float, currency: str = "RUB") -> str:
    sym = CURRENCY_SYMBOLS.get(currency, currency)
    return f"{value:,.2f} {sym}".replace(",", " ")